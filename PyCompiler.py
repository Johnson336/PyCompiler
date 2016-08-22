#! requires python 2.5+
# compiler.py - Scans weight sheets in a directory and spits out compiled data from them
# @Author - Cole Niermeyer
# @Date - 8/20/16 04:27:02 PM

import os, sys, openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment

"""
########### Change these to #############
######### match scrap template ##########
#########################################
"""
# INPUT

# ROW PASSED ITEM HEADERS ARE ON (EM, COMP, IH, FG)
PASSEDHEADERROW = 11
# COLUMN OF EM
EMCOL = 'S'
# COLUMN OF COMPONENTS
COMPCOL = 'T'
# COLUMN OF IN HOUSE
INHOUSECOL = 'U'
# COLUMN OF FG
FGCOL = 'V'
# STARTING COLUMN OF SCRAP
SCRAPSTARTCOL = 'C'
# ENDING COLUMN OF SCRAP
SCRAPENDCOL = 'R'
# ROW SCRAP ITEM HEADERS ARE ON (Servers, Laptops, Desktops, etc)
SCRAPHEADERROW = 11

"""
#####################################################
########### Change these to suit output #############
#####################################################
"""
# OUTPUT

# CHOOSE OUTPUT FILENAME
OUTFILE = 'Compiled Weight.xlsx'
# CHOOSE OUTPUT SHEET INSIDE OUTFILE
OUTSHEETNAME = 'autocompiled'

# CHOOSE PLACEMENT OF OUTPUT SHEET ITEMS
PASSEDOUTPUTROW = 12
EMOUTPUTCOL = 'D'
COMPOUTPUTCOL = 'E'
INHOUSEOUTPUTCOL = 'F'
FGOUTPUTCOL = 'G'
CPUOUTPUTCOL = 'I'
RAMOUTPUTCOL = 'K'
INOPOUTPUTCOL = 'H'
SCRAPOUTPUTHEADERROW = 11
SCRAPOUTPUTCOL = 'M'


"""
##############################################
####### DON'T CHANGE ANYTHING BELOW HERE######
##############################################
"""

files = []

emData = []
compData = []
inhouseData = []
fgData = []

scrapTitles = []
scrapItems = []
scrapCol = []

inopData = []
CPUData = []
RAMData = []

def main():
  global scrapCol
  
  print('Scanning ' + sys.path[0] + ' for files...')
  for f in os.listdir(sys.path[0]):
    if f.endswith(".xlsx") and f[:1].isdigit():
      files.append(f)
  print('Found ' + str(len(files)) + ' files, compiling...')
  
  for f in files:    
    print('Compiling ' + f + '...')
    book = openpyxl.load_workbook(f)
    sheet = book.get_sheet_by_name('eTest Weight')
  
    # get passed item data
    for row in range(PASSEDHEADERROW + 1, 75):
      # EM
      emCell = sheet[EMCOL + str(row)]
      if emCell.value:
        # and (emCell.fill.bgColor.rgb == '00000000')
        emData.append(emCell.value)
      # COMPONENTS
      compCell = sheet[COMPCOL + str(row)]
      if compCell.value:
        compData.append(compCell.value)
      # IN HOUSE
      ihCell = sheet[INHOUSECOL + str(row)]
      if ihCell.value:
        inhouseData.append(ihCell.value)
      # FG
      fgCell = sheet[FGCOL + str(row)]
      if fgCell.value:
        fgData.append(fgCell.value)
    
    # get scrap item data
    for col in range(column_index_from_string(SCRAPSTARTCOL), column_index_from_string(SCRAPENDCOL)+1):
      scrapTitle = sheet[get_column_letter(col) + str(SCRAPHEADERROW)].value
      if scrapTitle:
        found = False
        for row in range(SCRAPHEADERROW+1, sheet.max_row + 1):
          scrapCell = sheet[get_column_letter(col) + str(row)]
          if scrapCell.value:
            scrapCol.append(scrapCell.value)
            found = True
        if found:
          scrapItems.append(scrapCol)
          scrapCol = []
          scrapTitles.append(scrapTitle)
  
  # strip out inop, cpu and ram items
  strippedIndexes = []
  for i, v in enumerate(scrapTitles):
    if "inop" in v.lower():
      inopData.append(scrapTitles[i])
      inopData.extend(scrapItems[i])
      strippedIndexes.append(i)
    elif "cpu" in v.lower():
      CPUData.extend(scrapItems[i])
      strippedIndexes.append(i)
    elif "ram" in v.lower():
      RAMData.extend(scrapItems[i])
      strippedIndexes.append(i)
      
  for i in reversed(strippedIndexes):
    del scrapTitles[i]
    del scrapItems[i]
  
  #outbook = openpyxl.Workbook()
  outsheetindex = 0
  if OUTFILE in os.listdir(sys.path[0]):
    outsheetindex = 1
    outbook = openpyxl.load_workbook(OUTFILE)
  else:
    outbook = openpyxl.Workbook()
  outsheet = outbook.get_sheet_names()
  if OUTSHEETNAME in outsheet:
    outsheet = outbook.get_sheet_by_name(OUTSHEETNAME)
  else:
    outsheet = outbook.create_sheet(index=outsheetindex, title=OUTSHEETNAME)
    
  outsheet.merge_cells('D10:G10')
  outsheet['D10'] = 'PASSED ITEMS'
  outsheet['D10'].alignment = Alignment(horizontal='center')
  outsheet['D11'] = 'EM'
  outsheet['E11'] = 'COMPONENTS'
  outsheet['F11'] = 'In-House'
  outsheet['G11'] = 'FG'
  outsheet.merge_cells('M10:BB10')
  outsheet['M10'] = 'SCRAP ITEMS'
  outsheet['M10'].alignment = Alignment(horizontal='center')
  outsheet.merge_cells('I10:L10')
  outsheet['I10'] = 'COMPONENT SCRAP'
  outsheet['I10'].alignment = Alignment(horizontal='center')
  outsheet['I11'] = 'CPU'
  outsheet['K11'] = 'RAM'
  outsheet['H11'] = 'INOP'
  
  # output compiled data onto outsheetname
  for i, d in enumerate(emData):
    outsheet[EMOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(compData):
    outsheet[COMPOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(inhouseData):
    outsheet[INHOUSEOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(fgData):
    outsheet[FGOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(scrapTitles):
    outsheet[get_column_letter(column_index_from_string(SCRAPOUTPUTCOL)+i)+str(SCRAPOUTPUTHEADERROW)].value = d
  for a, b in enumerate(scrapItems):
    for k, v in enumerate(b):
      outsheet[get_column_letter(column_index_from_string(SCRAPOUTPUTCOL)+a)+str(k+SCRAPOUTPUTHEADERROW+1)].value = v
  for i, d in enumerate(inopData):
    #print(type(d))
    outsheet[INOPOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(CPUData):
    outsheet[CPUOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(RAMData):
    outsheet[RAMOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  
  
  outbook.save(OUTFILE)
  print("Compilation saved to '" + OUTSHEETNAME + "' sheet in '" + OUTFILE + "'.")

main()
os.system("pause")


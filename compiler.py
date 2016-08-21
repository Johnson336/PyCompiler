#! requires python 2.5+
# compiler.py - Scans weight sheets in a directory and spits out compiled data from them
# @Author - Cole Niermeyer
# @Date - 8/20/16 04:27:02 PM

import os, sys, openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment

"""
########### Change these to #############
######### match scrap template ##########
#########################################
"""
# INPUT

# ROW PASSED ITEM HEADERS ARE ON (EM, COMP, IH, FG)
PASSEDHEADERROW = 7
# COLUMN OF EM
EMCOL = 'T'
# COLUMN OF COMPONENTS
COMPCOL = 'U'
# COLUMN OF IN HOUSE
INHOUSECOL = 'V'
# COLUMN OF FG
FGCOL = 'W'
# STARTING COLUMN OF SCRAP
SCRAPSTARTCOL = 'C'
# ENDING COLUMN OF SCRAP
SCRAPENDCOL = 'S'
# ROW SCRAP ITEM HEADERS ARE ON (Servers, Laptops, Desktops, etc)
SCRAPHEADERROW = 7

"""
#####################################################
########### Change these to suit output #############
#####################################################
"""
# OUTPUT

# CHOOSE OUTPUT FILENAME
OUTFILE = 'Compiled Weight.xlsx'
# CHOOSE OUTPUT SHEET INSIDE OUTFILE
OUTSHEETNAME = 'autocompiled'

# CHOOSE PLACEMENT OF OUTPUT SHEET ITEMS
PASSEDOUTPUTROW = 10
EMOUTPUTCOL = 'C'
COMPOUTPUTCOL = 'D'
INHOUSEOUTPUTCOL = 'E'
FGOUTPUTCOL = 'F'
CPUOUTPUTCOL = 'G'
RAMOUTPUTCOL = 'H'
INOPOUTPUTCOL = 'I'
SCRAPOUTPUTHEADERROW = 9
SCRAPOUTPUTCOL = 'J'


"""
##############################################
####### DON'T CHANGE ANYTHING BELOW HERE######
##############################################
"""

files = []

emData = []
compData = []
inhouseData = []
fgData = []

scrapTitles = []
scrapItems = []
scrapCol = []

inopData = []
CPUData = []
RAMData = []

def main():
  global scrapCol
  print('Scanning ' + sys.path[0] + ' for files...')
  for f in os.listdir(sys.path[0]):
    if (f.endswith(".xlsx") or f.endswith(".xls")) and f[:1].isdigit():
      files.append(f)
  print('Found ' + str(len(files)) + ' files, compiling...')
  
  for f in files:
    print('Compiling ' + f + '...')
    book = openpyxl.load_workbook(f)
    sheet = book.get_sheet_by_name('Sheet1')
  
    # get passed item data
    for row in range(8, sheet.max_row + 1):
      # EM
      emCell = sheet[EMCOL + str(row)]
      if emCell.value and (emCell.fill.bgColor.rgb == '00000000'):
        emData.append(emCell.value)
      # COMPONENTS
      compCell = sheet[COMPCOL + str(row)]
      if compCell.value and (compCell.fill.bgColor.rgb == '00000000'):
        compData.append(compCell.value)
      # IN HOUSE
      ihCell = sheet[INHOUSECOL + str(row)]
      if ihCell.value and (ihCell.fill.bgColor.rgb == '00000000'):
        inhouseData.append(ihCell.value)
      # FG
      fgCell = sheet[FGCOL + str(row)]
      if fgCell.value and (fgCell.fill.bgColor.rgb == '00000000'):
        fgData.append(fgCell.value)
    
    # get scrap item data
    for col in range(column_index_from_string(SCRAPSTARTCOL), column_index_from_string(SCRAPENDCOL)+1):
      scrapTitle = sheet[get_column_letter(col) + str(SCRAPHEADERROW)].value
      if scrapTitle:
        found = False
        for row in range(SCRAPHEADERROW+1, sheet.max_row + 1):
          scrapCell = sheet[get_column_letter(col) + str(row)]
          if scrapCell.value and (scrapCell.fill.bgColor.rgb == '00000000'):
            scrapCol.append(scrapCell.value)
            found = True
        if found:
          scrapItems.append(scrapCol)
          scrapCol = []
          scrapTitles.append(scrapTitle)
  
  # strip out inop, cpu and ram items
  strippedIndexes = []
  for i, v in enumerate(scrapTitles):
    if "inop" in v.lower():
      inopData.append(scrapTitles[i])
      inopData.extend(scrapItems[i])
      strippedIndexes.append(i)
    elif "cpu" in v.lower():
      CPUData.extend(scrapItems[i])
      strippedIndexes.append(i)
    elif "ram" in v.lower():
      RAMData.extend(scrapItems[i])
      strippedIndexes.append(i)
      
  for i in reversed(strippedIndexes):
    del scrapTitles[i]
    del scrapItems[i]
  
  #outbook = openpyxl.Workbook()
  outsheetindex = 0
  if OUTFILE in os.listdir(sys.path[0]):
    outsheetindex = 1
    outbook = openpyxl.load_workbook(OUTFILE)
  else:
    outbook = openpyxl.Workbook()
  outsheet = outbook.get_sheet_names()
  if OUTSHEETNAME in outsheet:
    outsheet = outbook.get_sheet_by_name(OUTSHEETNAME)
  else:
    outsheet = outbook.create_sheet(index=outsheetindex, title=OUTSHEETNAME)
    
  outsheet.merge_cells('C8:F8')
  outsheet['C8'] = 'PASSED ITEMS'
  outsheet['C8'].alignment = Alignment(horizontal='center')
  outsheet['C9'] = 'EM'
  outsheet['D9'] = 'COMPONENTS'
  outsheet['E9'] = 'In-House'
  outsheet['F9'] = 'FG'
  outsheet.merge_cells('J8:AA8')
  outsheet['J8'] = 'SCRAP'
  outsheet['J8'].alignment = Alignment(horizontal='center')
  outsheet.merge_cells('G8:H8')
  outsheet['G8'] = 'COMPONENT SCRAP'
  outsheet['G8'].alignment = Alignment(horizontal='center')
  outsheet['G9'] = 'CPU'
  outsheet['H9'] = 'RAM'
  outsheet['I9'] = 'INOP'
  
  # output compiled data onto outsheetname
  for i, d in enumerate(emData):
    outsheet[EMOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(compData):
    outsheet[COMPOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(inhouseData):
    outsheet[INHOUSEOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(fgData):
    outsheet[FGOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(scrapTitles):
    outsheet[get_column_letter(column_index_from_string(SCRAPOUTPUTCOL)+i)+str(SCRAPOUTPUTHEADERROW)].value = d
  for a, b in enumerate(scrapItems):
    for k, v in enumerate(b):
      outsheet[get_column_letter(column_index_from_string(SCRAPOUTPUTCOL)+a)+str(k+SCRAPOUTPUTHEADERROW+1)].value = v
  for i, d in enumerate(inopData):
    #print(type(d))
    outsheet[INOPOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(CPUData):
    outsheet[CPUOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  for i, d in enumerate(RAMData):
    outsheet[RAMOUTPUTCOL + str(PASSEDOUTPUTROW+i)].value = d
  
  
  outbook.save(OUTFILE)
  print("Compilation saved to '" + OUTSHEETNAME + "' sheet in '" + OUTFILE + "'.")

main()
os.system("pause")



